import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.filters import AutoFilter
from datetime import date, timedelta

wb = openpyxl.Workbook()

# ── Colors ──────────────────────────────────────────────────
ORANGE      = "EA580C"
WHITE       = "FFFFFF"
HOT_RED     = "FECACA"
WARM_YEL    = "FEF08A"
COLD_GREY   = "E5E7EB"
GREEN_FILL  = "BBFABB"
RED_FILL    = "FECACA"
YEL_FILL    = "FEF08A"
HEADER_FONT = Font(bold=True, color=WHITE, size=11)
ORANGE_FILL = PatternFill("solid", fgColor=ORANGE)

def hdr(ws, col, row, val):
    c = ws.cell(row=row, column=col, value=val)
    c.fill  = ORANGE_FILL
    c.font  = HEADER_FONT
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c

def wa_link(phone_raw):
    if not phone_raw:
        return ""
    d = "".join(ch for ch in str(phone_raw) if ch.isdigit())
    if d.startswith("0"):
        d = d[1:]
    if not d.startswith("54"):
        d = "54" + d
    return f"https://wa.me/{d}"

# ════════════════════════════════════════════════════════════
# HOJA 1 — Prospectos
# ════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Prospectos"

headers = [
    "ID","🔥 Prioridad","Score","Nombre","Especialidad","Ciudad / Barrio",
    "Teléfono","Instagram","WhatsApp (link)","URL Perfil","Notas iniciales",
    "📞 Fecha Llamada 1","Estado Llamada 1","📲 Fecha Demo enviada",
    "👀 Probó la demo?","🚦 Señal de interés",
    "📞 Fecha Llamada 2","💰 Cerrado?","Notas"
]
for i, h in enumerate(headers, 1):
    hdr(ws1, i, 1, h)

ws1.freeze_panes = "A2"
ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

COL_W = [6,10,7,26,24,22,16,18,36,40,30,16,18,18,14,18,16,12,30]
for i, w in enumerate(COL_W, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.row_dimensions[1].height = 32

# ── Lead data ────────────────────────────────────────────────
# Fields: nombre, especialidad, ciudad_barrio, telefono, instagram, url_perfil, notas
# Score computed automatically
LEADS_RAW = [
    # ─ CÓRDOBA – Cales Network (tienen WA 3512308121) ─────
    ("Lic. María Constanza Alaye",   "Psicología clínica / EMDR",    "Córdoba Capital",       "3512308121","","https://www.doctoraliar.com/maria-constanza-alaye/psicologo/cordoba-capital","114 reseñas Doctoralia. Red Cales."),
    ("Lic. Vilma Barrios",           "Psicoanálisis / Clínica",      "Córdoba Capital",       "3512308121","","https://www.doctoraliar.com/vilma-barrios/psicologo/cordoba-capital","48 reseñas. Red Cales."),
    ("Lic. Araceli Mañanes",         "Psicología clínica",           "Córdoba Capital",       "3512308121","","https://www.doctoraliar.com/araceli-mananes/psicologo/cordoba-capital","38 reseñas. Red Cales."),
    ("Lic. Silvina Jimenez",         "Psicoanálisis / Clínica",      "Córdoba Capital",       "3512308121","","https://www.doctoraliar.com/silvina-jimenez/psicologo/cordoba-capital","36 reseñas. Red Cales."),
    ("Lic. Magalí Langone",          "Gestalt / Clínica",            "Córdoba Capital",       "3512308121","","https://www.doctoraliar.com/magali-langone/psicologo/cordoba-capital","26 reseñas. Red Cales."),
    ("Lic. Daniela Agüero",          "Psicoanálisis",                "Córdoba Capital",       "3512308121","","https://www.doctoraliar.com/daniela-aguero/psicologo/cordoba-capital","23 reseñas. Red Cales."),
    ("Lic. Daniela Sommer",          "Psicología clínica",           "Cerro de las Rosas, CBA","3512308121","","https://cales.com.ar/profesional/lic-daniela-soledad-sommer/","Red Cales WA"),
    ("Lic. Martina Nobile Giusto",   "Psicología clínica",           "Cerro de las Rosas, CBA","3512308121","","https://cales.com.ar/profesional/lic-martina-nobile-giusto/","Red Cales WA"),
    ("Lic. Silvana Oviedo",          "Psicología clínica",           "Cerro de las Rosas, CBA","3512308121","","https://cales.com.ar/profesional/lic-silvana-oviedo/","Red Cales WA"),
    ("Lic. Martin Diego Garro",      "Psicología clínica",           "Nueva Córdoba",          "3512308121","","https://cales.com.ar/profesional/lic-martin-diego-garro/","Red Cales WA"),
    ("Lic. Carla Sigifredo",         "Psicología clínica",           "Nueva Córdoba / Alberdi","3512308121","","https://cales.com.ar/profesional/lic-carla-sigifredo/","Red Cales WA"),
    ("Lic. Fabiana Paolini",         "Psicología clínica",           "Alberdi, Córdoba",       "3512308121","","https://cales.com.ar/profesional/lic-fabiana-maria-paolini/","Red Cales WA"),
    ("Lic. Florencia Cárdenas",      "Psicología clínica",           "General Paz, Córdoba",   "3512308121","","https://cales.com.ar/profesional/lic-florencia-cardenas/","Red Cales WA"),
    ("Lic. Valeria Pinotti",         "Psicología clínica",           "Córdoba Capital Centro", "3512308121","","https://cales.com.ar/profesional/lic-valeria-pinotti/","Red Cales WA"),
    ("Lic. Verónica Martinez",       "Psicología clínica",           "Córdoba Capital Centro", "3512308121","","https://cales.com.ar/profesional/lic-martinez-veronica-andrea/","Red Cales WA"),
    # ─ CÓRDOBA – independientes con teléfono ──────────────
    ("Lic. Cecilia A. Pegoraro",     "Psicología clínica",           "Córdoba Capital",        "0351 152832844","","","MP 7601. Tel propio."),
    ("Eugenia Duarte",               "Gestalt / Clínica",            "Córdoba Capital",        "","","https://psicologa.com.ar/","Online + presencial. 23 años exp."),
    # ─ PALERMO CABA – Nutricionistas ──────────────────────
    ("Dra. Maria Alejandra Moisello","Nutricionista",                "Palermo, CABA",          "","","https://www.doctoraliar.com/maria-alejandra-moisello/nutricionista/capital-federal","205 reseñas Doctoralia. Paraguay 2468 P2B."),
    ("Lic. Margarita Grundnig",      "Nutrición / Coaching nutric.", "Palermo-Botánico, CABA", "","","https://www.doctoraliar.com/margarita-grundnig/nutricionista/capital-federal","36 reseñas. Armenia 2387."),
    ("Dra. Cecilia Trujillo",        "Nutrición / Medicina integral.","Palermo, CABA",          "","","https://www.doctoraliar.com/cecilia-trujillo/medico-general-y-familiar-nutricionista/capital-federal","16 reseñas. Pueyrredón 1705."),
    # ─ PALERMO CABA – Psicólogos ──────────────────────────
    ("Lic. Antonella Galanti",       "TCC / Terapias contextuales",  "Recoleta/Palermo, CABA", "","","https://www.doctoraliar.com/antonella-galanti/psicologo/capital-federal","77 reseñas. Av. Libertador 1428."),
    ("Lic. Janina Salatti",          "Psicología laboral / clínica", "Palermo, CABA",          "","","https://www.doctoraliar.com/janina-salatti/psicologo/rio-gallegos","79 reseñas. Paraguay 3042."),
    ("Lic. Celeste Diaz",            "Psicología forense / Ansiedad","Palermo, CABA",          "","","https://www.doctoraliar.com/celeste-diaz/psicologo/benavidez","51 reseñas. Av. Santa Fe."),
    ("Estrin Silvia Noemi",          "Ansiedad / Terapia pareja",    "Palermo, CABA",          "","","https://www.mundopsicologos.com.ar/centros/estrin-silvia-noemi","Gurruchaga 506. Turnos online."),
    ("Lic. Guido Pazos",             "TCC / Clínica",                "Palermo, CABA",          "","","https://www.mundopsicologos.com.ar/centros/lic-guido-pazos","Aráoz 2268."),
    ("Lic. Gonzalo Shalom",          "Psicología clínica (UBA)",     "Palermo, CABA",          "","","https://www.mundopsicologos.com.ar/centros/lic-gonzalo-shalom","MN 70833. Presencial y online."),
    ("Karina Matzkin",               "Ansiedad / Depresión",         "Palermo, CABA",          "","","https://www.mundopsicologos.com.ar/centros/karina-matzkin","Paraguay 3636."),
    ("Lic. Martina Michelini",       "TCC / Autoestima",             "Palermo, CABA",          "","","https://www.mundopsicologos.com.ar/centros/lic-martina-michelini","Presencial y online."),
    ("Lic. Lautaro Cisneros",        "Ansiedad / Terapia pareja",    "Palermo, CABA",          "","","https://www.mundopsicologos.com.ar/centros/lic-lautaro-cisneros","Paraguay 4176."),
    ("Gabriela Fernández Ortiz",     "Ansiedad / Psic. infantil",    "Palermo, CABA",          "","","https://www.mundopsicologos.com.ar/centros/gabriela-fernandez-ortiz","Mansilla 3333."),
    ("Victoria Henao",               "Especialista ansiedad",        "Palermo, CABA",          "","","https://psicologosbuenosaires.com/psicologa-victoria-henao-esp-ansiedad-672","Zapata 260."),
    ("Diego A. Valdivia",            "TCC",                          "Palermo, CABA",          "","","https://psicologosbuenosaires.com/diego-a-valdivia-psicologo-tcc--1429","Araoz 2850. $23,000/ses."),
    ("Lic. Federico Plá",            "TCC",                          "Palermo, CABA",          "","","https://psicologosbuenosaires.com/lic-federico-pla-psicologo-cognitivo-conductual-317","Sánchez de Bustamante."),
    ("Valentina Barrios",            "TCC / Cognitivo conductual",   "Palermo, CABA",          "","","https://psicologosbuenosaires.com/lic-valentina-barrios-psicologa-cognitivo-conductual-340",""),
    ("Sofía Duarte",                 "Psicología clínica / TCC",     "Palermo, CABA",          "","","https://psicologosbuenosaires.com/especialista-en-psicologia-clinica-uba-terapeuta-cognitivo-conductual-1531","Migueletes 825."),
    ("Jonathan Petrone",             "TCC",                          "Palermo, CABA",          "","","https://psicologosbuenosaires.com/lic-jonathan-petrone-psicologo-cognitivo-conductual-1929","Plaza Italia."),
    ("Graciela Mónica Guidi",        "Ansiedad / Terapia pareja",    "Palermo, CABA",          "","","https://www.mundopsicologos.com.ar/centros/graciela-monica-guidi","Malabia 2363."),
    ("Melina Oundjian",              "Ansiedad / Psic. infantil",    "Palermo, CABA",          "","","https://www.mundopsicologos.com.ar/centros/melina-oundjian","Ombú 2994."),
    ("Mercedes De Nicolás",          "Psicoanálisis",                "Palermo, CABA",          "","","https://psicologosbuenosaires.com/mercedes-de-nicolas-psicologa-psicoanalista-297","Gurruchaga 2100."),
    # ─ RECOLETA CABA ──────────────────────────────────────
    ("Lic. Florencia Larrarte",      "TCC / Terapia pareja",        "Recoleta, CABA",          "","","https://www.mundopsicologos.com.ar/centros/lic-florencia-larrarte","Gallo 1071."),
    ("Gustavo Rimoli",               "TCC / Clínica",               "Recoleta, CABA",          "","","https://www.mundopsicologos.com.ar/centros/gustavo-rimoli","Las Heras 2050."),
    ("Margarita Brassara",           "Ansiedad / Psic. adolesc.",   "Recoleta, CABA",          "","","https://www.mundopsicologos.com.ar/centros/margarita-brassara","Pueyrredón 1005."),
    ("Lic. Javier Lagorio",          "Ansiedad / Psic. infantil",   "Recoleta, CABA",          "","","https://www.mundopsicologos.com.ar/centros/lic-javier-ignacio-lagorio","Anchorena 1484."),
    ("María Paula de Veyga",         "TCC / Psic. infantil",        "Recoleta, CABA",          "","","https://www.mundopsicologos.com.ar/centros/lic-maria-paula-de-veyga","Austria 2122."),
    ("Bárbara Ballesteros",          "TCC / Ansiedad",              "Palermo/Recoleta, CABA",  "","","https://www.mundopsicologos.com.ar/centros/barbara-ballesteros","Av. Santa Fe y Callao. $42k."),
    # ─ BELGRANO CABA ──────────────────────────────────────
    ("Lic. Jonatan Sallustio",       "TCC / Psicoanálisis",         "Belgrano, CABA",           "","","https://www.doctoraliar.com/jonatan-sallustio/psicologo-psicoanalista/capital-federal","75 reseñas. Av. Melián 2067."),
    ("Lic. María Sol Cátera",        "Psicología clínica",          "Belgrano, CABA",           "","","https://www.doctoraliar.com/maria-sol-catera/psicologo/capital-federal","46 reseñas. José Hernández 2729."),
    ("Lic. Gabriela Cirillo",        "Psicoterapia individual",     "Belgrano, CABA",           "","","https://www.doctoraliar.com/gabriela-cirillo/psicologo/capital-federal","26 reseñas. Av. Luis María Campos."),
    ("Lic. Diego Pimentel",          "Neuropsicología / EMDR",      "Belgrano, CABA",           "","","https://www.doctoraliar.com/diego-hernan-pimentel/psicologo/capital-federal","37 reseñas. Av. Cabildo 3062."),
    ("Lic. Gabriela Rozados",        "TCC / Clínica",               "Belgrano, CABA",           "","","https://www.doctoraliar.com/gabriela-rozados/psicologo/belgrano","10 reseñas. Cramer 2734."),
    ("Muriel Noguera",               "Psicoanálisis / Ansiedad",    "Belgrano, CABA",           "","","https://www.mundopsicologos.com.ar/centros/muriel-noguera","Av. Monroe."),
    ("Sol Mouriño Marzoa",           "Ansiedad / Psic. infantil",   "Belgrano, CABA",           "","","https://www.mundopsicologos.com.ar/centros/sol-mourino-marzoa","Av. Cabildo y Lacroze."),
    ("Lic. Carolina Ramos",          "TCC / Terapia pareja",        "Belgrano, CABA",           "","","https://www.mundopsicologos.com.ar/centros/lic-carolina-ramos","Jose Hernandez 1955. Online."),
    ("Lic. Marcela Maidana",         "Ansiedad / Pánico",           "Belgrano, CABA",           "","","https://www.mundopsicologos.com.ar/centros/lic-marcela-maidana","Roosevelt 2400. Online."),
    ("Lic. Laura Sussi",             "TCC / Terapia pareja",        "Belgrano, CABA",           "","","https://www.mundopsicologos.com.ar/centros/lic-laura-sussi","Cabildo."),
    ("Lic. Vivián Castro Cabral",    "TCC / Duelo",                 "Belgrano, CABA",           "","","https://www.mundopsicologos.com.ar/centros/lic-vivian-fabiana-castro-cabral","11 de Septiembre 1651."),
    ("Lic. Mirta Seguí",             "Ansiedad / Pánico",           "Belgrano, CABA",           "","","https://www.mundopsicologos.com.ar/centros/lic-mirta-haydee-segui","Cuba 1871."),
    ("Lic. Belinda Tancredi",        "TCC / Duelo",                 "Belgrano, CABA",           "","","https://www.mundopsicologos.com.ar/centros/lic-belinda-tancredi","Ciudad de la Paz 2139."),
    ("Lic. Sandra Gundin",           "Terapia familiar / Clínica",  "Belgrano, CABA",           "","","https://www.mundopsicologos.com.ar/centros/lic-sandra-gundin","La Pampa 2477."),
    # ─ NUTRICIONISTAS CÓRDOBA ────────────────────────────
    ("Lic. Vanesa Birri",            "Nutricionista",               "Córdoba Capital",           "","","http://www.nutricionistaencordoba.com/","Bv. San Juan 340. OSDE/SANCOR."),
    ("Nutriser Centro Nutrición",    "Nutricionista (centro)",      "Córdoba Capital",           "","","https://centronutriser.com/","Consultas presenciales y online."),
]

def score_lead(nombre, esp, ciudad, tel, url, notas):
    s = 0
    if tel:                                  s += 3
    if url:                                  s += 2
    if "doctoralia" in url.lower() or "mundopsicologos" in url.lower(): s += 2
    combined = (esp + notas + url).lower()
    if "online" in combined or "particular" in combined or "turnos" in combined: s += 2
    if any(k in combined for k in ["tcc","cognitivo","clínica","clinica","emdr","gestalt"]): s += 1
    return min(s, 10)

def priority(s):
    if s >= 8:  return "🔥 HOT"
    if s >= 5:  return "⚡ WARM"
    return "🥶 COLD"

scored = []
for idx, row in enumerate(LEADS_RAW, 1):
    nombre, esp, ciudad, tel, ig, url, notas = row
    s = score_lead(nombre, esp, ciudad, tel, url, notas)
    scored.append((s, idx, nombre, esp, ciudad, tel, ig, url, notas))

scored.sort(key=lambda x: -x[0])

HOT_F  = PatternFill("solid", fgColor=HOT_RED)
WARM_F = PatternFill("solid", fgColor=WARM_YEL)
COLD_F = PatternFill("solid", fgColor=COLD_GREY)

for seq, (s, _orig, nombre, esp, ciudad, tel, ig, url, notas) in enumerate(scored, 1):
    p = priority(s)
    wa = wa_link(tel) if tel else ""
    row_data = [seq, p, s, nombre, esp, ciudad, tel, ig, wa, url, notas,
                "", "", "", "", "", "", "", ""]
    r = seq + 1
    for col, val in enumerate(row_data, 1):
        c = ws1.cell(row=r, column=col, value=val)
        c.alignment = Alignment(vertical="center", wrap_text=False)
    prio_cell = ws1.cell(row=r, column=2)
    if "HOT"  in p: prio_cell.fill = HOT_F
    elif "WARM" in p: prio_cell.fill = WARM_F
    else:             prio_cell.fill = COLD_F

# Conditional formatting col R (18) for Cerrado
from openpyxl.formatting.rule import CellIsRule
green_f = PatternFill("solid", fgColor="BBFABB")
red_f   = PatternFill("solid", fgColor=HOT_RED)
yel_f   = PatternFill("solid", fgColor=WARM_YEL)
last_row = len(scored) + 1
ws1.conditional_formatting.add(
    f"R2:R{last_row}",
    CellIsRule(operator="equal", formula=['"Sí"'],   fill=green_f))
ws1.conditional_formatting.add(
    f"R2:R{last_row}",
    CellIsRule(operator="equal", formula=['"No"'],   fill=red_f))
ws1.conditional_formatting.add(
    f"R2:R{last_row}",
    CellIsRule(operator="equal", formula=['"Lo piensa"'], fill=yel_f))

# ════════════════════════════════════════════════════════════
# HOJA 2 — Dashboard
# ════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("📊 Dashboard")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 18

def dash_label(ws, row, label, formula, fmt=None):
    lbl = ws.cell(row=row, column=1, value=label)
    lbl.font = Font(bold=True, size=12)
    lbl.alignment = Alignment(horizontal="right")
    val = ws.cell(row=row, column=2, value=formula)
    val.font = Font(bold=True, size=18, color="EA580C")
    val.alignment = Alignment(horizontal="center")
    if fmt:
        val.number_format = fmt
    return val

ws2.cell(row=1, column=1, value="📊 LUMIO CRM — DASHBOARD").font = Font(bold=True, size=16, color=ORANGE)
dash_label(ws2, 2,  "Total leads",          "=COUNTA(Prospectos!D:D)-1")
dash_label(ws2, 3,  "Llamadas realizadas",  "=COUNTA(Prospectos!L:L)-1")
dash_label(ws2, 4,  "Demos enviadas",       "=COUNTA(Prospectos!N:N)-1")
dash_label(ws2, 5,  "Cerrados 💰",          "=COUNTIF(Prospectos!R:R,\"Sí\")")
dash_label(ws2, 6,  "Tasa demo",            "=IFERROR(B4/B3,0)", "0%")
dash_label(ws2, 7,  "Tasa cierre",          "=IFERROR(B5/B4,0)", "0%")
dash_label(ws2, 8,  "Setup cobrado (USD)",  "=B5*200", '"USD "#,##0')
dash_label(ws2, 9,  "MRR comprometido",     "=B5*99",  '"USD "#,##0')

ws2.cell(row=11, column=1, value="— Por Ciudad —").font = Font(bold=True, size=12)
cities = ["Palermo, CABA","Recoleta, CABA","Belgrano, CABA","Córdoba Capital","Nueva Córdoba","Cerro de las Rosas, CBA"]
for i, city in enumerate(cities, 12):
    ws2.cell(row=i, column=1, value=city).font = Font(size=11)
    ws2.cell(row=i, column=2,
             value=f'=COUNTIF(Prospectos!F:F,"*{city.split(",")[0]}*")').font = Font(bold=True, size=13)

ws2.cell(row=19, column=1, value="— Por Prioridad —").font = Font(bold=True, size=12)
for i, (label, kw) in enumerate([("🔥 HOT","HOT"),("⚡ WARM","WARM"),("🥶 COLD","COLD")], 20):
    ws2.cell(row=i, column=1, value=label).font = Font(size=11)
    ws2.cell(row=i, column=2, value=f'=COUNTIF(Prospectos!B:B,"*{kw}*")').font = Font(bold=True, size=13)

# ════════════════════════════════════════════════════════════
# HOJA 3 — Guion
# ════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("🎙️ Guion + Mensajes")
ws3.column_dimensions["A"].width = 90
ws3.sheet_view.showGridLines = False

script_lines = [
("📞 GUION LLAMADA 1 (memorizar)", True, 16),
("","",11),
("APERTURA (8 seg):", True, 12),
('"Hola, ¿hablo con [consultorio]? Soy Fran de Lumio. Te robo 30 segundos y si no te interesa cortamos sin problema, ¿va?"', False, 11),
("","",11),
("GANCHO (15 seg):", True, 12),
('"Trabajamos con consultorios de psicología en Argentina automatizando atención por WhatsApp con IA. El bot toma turnos, confirma sesiones y reduce ausencias 50%. ¿Hoy en tu consultorio, quién responde los WhatsApp?"', False, 11),
("","",11),
("PIVOT a demo:", True, 12),
('"Mirá, lo mejor es que lo veas funcionando. Te paso un asistente en vivo para que lo pruebes como paciente. ¿A qué número te lo mando?"', False, 11),
("","",11),
("─"*60,"",11),
("📲 MENSAJE POST-LLAMADA (copiar y pegar)", True, 14),
("","",11),
("Hola [NOMBRE]! Soy Fran de Lumio, hablamos recién.", False, 11),
("Acá te dejo el asistente para que lo pruebes como paciente.", False, 11),
("📲 3535109880", False, 12),
("Probá pedirle turno, preguntarle precios, obras sociales — lo que harías como paciente.", False, 11),
("Cuando armemos el tuyo va a tener TU nombre, TUS servicios y TUS precios.", False, 11),
("Cualquier cosa me avisás!", False, 11),
("","",11),
("─"*60,"",11),
("🔁 MENSAJE TANTEO 48HS (si no responde)", True, 14),
("","",11),
("[NOMBRE]! ¿Pudiste probar el asistente?", False, 11),
("Tip: probalo de noche o un sábado — es justo el horario en que más mensajes recibís y no podés responder.", False, 11),
("Cualquier cosa me decís!", False, 11),
("","",11),
("─"*60,"",11),
("🚨 MANEJO DE OBJECIONES", True, 14),
("","",11),
('"¿Cuánto sale?" → "Arranca en USD 99/mes con setup único. Pero antes de hablar precios, probá el bot. Si no te genera valor, el precio es irrelevante."', False, 11),
("","",11),
('"No me interesa" → "Entiendo. Una pregunta: ¿es porque ya tenés un sistema, o porque hoy no es prioridad?"', False, 11),
("","",11),
('"Ya tengo Doctoralia" → "Eso te facilita. Pero Doctoralia no te atiende los WhatsApp ni confirma sesiones. ¿Cuántos pacientes te faltan sin avisar por semana?"', False, 11),
]

for r, (text, bold, size) in enumerate(script_lines, 1):
    c = ws3.cell(row=r, column=1, value=text)
    c.font = Font(bold=bool(bold), size=size, name="Arial")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws3.row_dimensions[r].height = 20 if not text else 28

# ════════════════════════════════════════════════════════════
# HOJA 4 — Métricas diarias
# ════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("📈 Métricas diarias")
met_headers = ["Fecha","Llamadas marcadas","Atendieron","Aceptaron demo",
               "Demos enviadas","Llamada 2 hecha","Cerrados","Notas día"]
for i, h in enumerate(met_headers, 1):
    c = ws4.cell(row=1, column=i, value=h)
    c.fill  = ORANGE_FILL
    c.font  = HEADER_FONT
    c.alignment = Alignment(horizontal="center")
    ws4.column_dimensions[get_column_letter(i)].width = 18

today = date.today()
for d in range(14):
    r = d + 2
    ws4.cell(row=r, column=1, value=today + timedelta(days=d+1)).number_format = "DD/MM/YYYY"
    for col in range(2, 8):
        ws4.cell(row=r, column=col, value=0)

tot_row = 17
ws4.cell(row=tot_row, column=1, value="TOTALES").font = Font(bold=True)
for col in range(2, 8):
    letter = get_column_letter(col)
    ws4.cell(row=tot_row, column=col,
             value=f"=SUM({letter}2:{letter}15)").font = Font(bold=True)

rate_row = 18
ws4.cell(row=rate_row, column=1, value="Tasas").font = Font(bold=True, italic=True)
ws4.cell(row=rate_row, column=3, value="=IFERROR(C17/B17,0)").number_format = "0%"
ws4.cell(row=rate_row, column=4, value="=IFERROR(D17/C17,0)").number_format = "0%"
ws4.cell(row=rate_row, column=5, value="=IFERROR(E17/D17,0)").number_format = "0%"
ws4.cell(row=rate_row, column=7, value="=IFERROR(G17/E17,0)").number_format = "0%"

# ════════════════════════════════════════════════════════════
# HOJA 5 — Cómo crecer el CRM
# ════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("📚 Cómo crecer el CRM")
ws5.column_dimensions["A"].width = 90
ws5.sheet_view.showGridLines = False

growth_lines = [
("📚 CÓMO CRECER ESTE CRM DESDE EL CELU", True, 16),
("","",11),
("LISTA INICIAL: ~60 leads sembrados automáticamente.", False, 11),
("META: llegar a 300 leads en 1 semana.", False, 11),
("","",11),
("CÓMO AGREGAR LEADS DESDE EL CELU (10 min/día):", True, 13),
("","",11),
("1. Abrí Instagram", False, 11),
("2. Buscá hashtags: #psicologapalermo #psicologacordoba #nutricionistabuenosaires #psicologabsas", False, 11),
("3. Por cada perfil interesante:", False, 11),
("   • Copiá el @ del IG", False, 11),
("   • Mirá si tiene WhatsApp en bio (anotá número)", False, 11),
("   • Agregá UNA FILA al final de la hoja Prospectos", False, 11),
("   • Mínimo: nombre + IG + ciudad", False, 11),
("4. 10 perfiles por día = 70 nuevos leads/semana", False, 11),
("","",11),
("FUENTES PARA SEGUIR LLENANDO:", True, 13),
("• doctoraliar.com/psicologo/[barrio]", False, 11),
("• mundopsicologos.com.ar/centros/[barrio]", False, 11),
("• Instagram (hashtags + ubicación)", False, 11),
("• Google Maps: \"psicólogo [barrio]\"", False, 11),
("• Colegio de Psicólogos de Córdoba (cppc.org.ar — matriculados públicos)", False, 11),
("• psicologosbuenosaires.com/[barrio]", False, 11),
("","",11),
("CRITERIO DE BUEN LEAD:", True, 13),
("✅ Tiene teléfono o IG con WhatsApp", False, 11),
("✅ Consultorio activo (posts recientes, reseñas)", False, 11),
("✅ Especialidad clínica (no demasiado nicho infantil)", False, 11),
("✅ Precio sesión USD 25-50 (sweet spot del bot)", False, 11),
("","",11),
("SCRIPT RÁPIDO PARA LLENAR FILA (copiar en notas del celu):", True, 13),
("Nombre | Esp | Ciudad | Tel | IG | URL | Notas", False, 11),
("Score: +3 tel, +2 URL, +2 Doctoralia/MundoPsi, +2 online/particular, +1 TCC/clínica", False, 11),
("HOT ≥8 | WARM 5-7 | COLD <5", False, 11),
]

for r, (text, bold, size) in enumerate(growth_lines, 1):
    c = ws5.cell(row=r, column=1, value=text)
    c.font = Font(bold=bool(bold), size=size, name="Arial")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws5.row_dimensions[r].height = 22

# ── Save ────────────────────────────────────────────────────
path = "/home/user/synergy-solutions/Lumio-CRM-Prospectos.xlsx"
wb.save(path)
print(f"SAVED: {path}")
print(f"LEADS: {len(scored)}")

# Summary
hot  = [x for x in scored if x[0] >= 8]
warm = [x for x in scored if 5 <= x[0] < 8]
cold = [x for x in scored if x[0] < 5]
print(f"HOT={len(hot)} WARM={len(warm)} COLD={len(cold)}")

from collections import Counter
cities = Counter(x[4].split(",")[0].strip() for x in scored)
print("TOP CIUDADES:", cities.most_common(4))

print("\nTOP 5 HOT:")
for s, _i, nombre, esp, ciudad, tel, ig, url, notas in hot[:5]:
    print(f"  [{s}] {nombre} | {ciudad} | tel:{tel or '-'} | {url[:50]}")
